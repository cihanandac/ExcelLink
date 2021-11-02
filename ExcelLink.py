import pandas as panda
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import shutil
import os.path
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog

"""This program is created to serve the need of giving automatic links of photos to a file, it will 
search a column of your choice sheet by sheet and if it finds any match in photo pools location, link 
will be given to cell of your choice.
Don't seperate the photo pool folder with the excel afterwards.

Feel free to use and share.
"""

print("Welcome to ExcelLink") 
print("You will be asked to choose the Excel file first and then")
print("the directories of the photos for each sheet.")
print(" "* 900)
#input('Press any key to continue')


root = tk.Tk()
root.withdraw()

#This is the excel file you want to add the links.
file_path = filedialog.askopenfilename(title="Choose the excel file")
print(file_path)
file = panda.ExcelFile(file_path)
sheets = file.sheet_names
wb = load_workbook(file_path)

#This is the directory which contains the photos that we will give links to.
directory_path = filedialog.askdirectory(title="Choose the location of the photo pool")

#iterating through the excel file for sheets
for sheet in sheets:
    print("Now working on the "+ sheet+ " sheet")
    page=file.parse(sheet)
    lenght, widht = page.shape
    print(lenght)
    ws = wb[sheet]
    sheet_path = directory_path + "/"+ sheet


    #iterating through the sheet for items
    for i in range(1, lenght):
        photo_check = page['Inv. No.'][i]
  
        #checking if there is a match
        #The algorithm for searching the filename is created according the need of the developer.
        
        for filename in os.listdir(sheet_path):
            
            #If your photo's name and the cell have the same name simply delete the codes until two lines above the Eureka part.
            first_line=0
            for k in range(0,len(filename)):
                if filename[k] == "_":
                    if first_line==0:
                        first_line=1
                    else:
                        second_line=0
                        for j in range(0,len(filename)):
                            if filename[j] == "_" or filename[j]== ".":
                                if second_line ==0 or second_line == 1:
                                    second_line = second_line + 1
                            
                                elif second_line ==2:
                                    shm_number = "SHM "+ filename[k+1:j]
                                    
                                    #This means that the photo's name is the same with the cell.
                                    if shm_number == photo_check:
                                        print("Eureka!!")
                                        
                                        #Feel free to change the column that you want to add the link.
                                        ws.cell(row=i+2, column=2).value = '=HYPERLINK("{}", "{}")'.format("Object_Photo/"+os.path.basename(directory_path)+"/"+sheet+"/"+ filename, "file")
                                        
                                        
                                        """ if you have the thumbnails for the photos you can also add them to another cell with following codes
                                        if os.path.isfile('Thumbnails/'+filename):
                                            img = openpyxl.drawing.image.Image('Thumbnails/'+filename+".jpg")
                                            img.anchor = "B"+str(i+2)
                                            ws.add_image(img)
                                        
                                        """

wb.save(file_path)


                                        


