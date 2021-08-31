import pandas as panda
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import shutil
import os.path
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog

print(" "* 2900)
print("Welcome to ExcelLink") 
print("This program is written for the purpose of linking excel cells")
print("with the related photos.") 
print("        ")
print("Feel free to use and share.")
print("Made by Cihan Andac")
print(" "* 1000)
print("You will be asked to choose the Excel file first and then")
print("the directories of the photos for each sheet.")
print(" "* 900)
#input('Press any key to continue')


root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(title="Choose the excel file")
print(file_path)


file = panda.ExcelFile(file_path)
sheets = file.sheet_names

wb = load_workbook(file_path)

directory_path = filedialog.askdirectory(title="Choose the location of the photo pool")



#iterating through the excel file for sheets
for sheet in sheets:
    print("Now working on the "+ sheet+ " sheet")
    page=file.parse(sheet)
    lenght, widht = page.shape
    print(lenght)
    ws = wb[sheet]
    sheet_path = directory_path + "/"+ sheet


    #iterating through the sheet for items
    for i in range(1, lenght):
        photo_check = page['Inv. No.'][i]
  
        #checking if there is a match
        
        for filename in os.listdir(sheet_path):

            
            first_line=0
            
            for k in range(0,len(filename)):
                if filename[k] == "_":
                    if first_line==0:
                        first_line=1
                    else:
                        second_line=0
                        for j in range(0,len(filename)):
                            if filename[j] == "_" or filename[j]== ".":
                                if second_line ==0 or second_line == 1:
                                    second_line = second_line + 1

                            
                                elif second_line ==2:
                                    shm_number = "SHM "+ filename[k+1:j]
                                    

                                    if shm_number == photo_check:
                                        print("eureka!!")
                                        if os.path.isfile('Thumbnails/'+filename):
                                            img = openpyxl.drawing.image.Image('Thumbnails/'+filename+".jpg")
                                            img.anchor = "B"+str(i+2)
                                            ws.cell(row=i+2, column=2).value = '=HYPERLINK("{}", "{}")'.format("Object_Photo/"+os.path.basename(directory_path)+"/"+sheet+"/"+ filename, "file")
                                            ws.add_image(img)
                                        else:
                                            ws.cell(row=i+2, column=2).value = '=HYPERLINK("{}", "{}")'.format("Object_Photo/"+os.path.basename(directory_path)+"/"+sheet+"/"+ filename, "file")
                                        
                                    
                                        
                                                                            
        

                    
wb.save(file_path)


                                        


